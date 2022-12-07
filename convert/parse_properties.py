import datetime
import json
import time

import openpyxl

"""
将json的属性解析为列
"""
# 暗网
dark_net = dict()
# 毒品
drug = dict()
# 赌博
gamble = dict()
# 恶意地址
malicious_address = dict()
# 洗钱
money_laundering = dict()
# 钓鱼诈骗
phishing_fraud = dict()
# 色情
pornography = dict()
# 传销
pyramid = dict()
# 诈骗
swindle = dict()
# 偷盗
steal = dict()
# 黑客
hacker = dict()
# 四方
square = dict()
# 混币合约
mix_currency_contract = dict()
# 异常中心化地址
exception_centralization_address = dict()


def init():
    """
    初始化属性名称
    :return:
    """
    dark_net["note"] = "备注"

    drug["angecy_name"] = "中介名称"
    drug["effect"] = "地址作用"
    drug["note"] = "注释"

    gamble["platform_name"] = "平台名称"
    gamble["platform_website"] = "平台网址"
    gamble["address_effect"] = "地址作用"
    gamble["address_type"] = "地址类型"
    gamble["sifang"] = "是否为四方"
    gamble["note"] = "备注"

    malicious_address["note"] = "备注"

    money_laundering["effect"] = "作用"
    money_laundering["note"] = "备注"

    phishing_fraud["note"] = "备注"
    phishing_fraud["address_type"] = "地址类型"
    phishing_fraud["address_effect"] = "地址作用"

    pornography["note"] = "备注"

    pyramid["platform_name"] = "平台名称"
    pyramid["platform_website"] = "平台网址"
    pyramid["address_effect"] = "地址作用"
    pyramid["address_type"] = "地址类型"
    pyramid["sifang"] = "是否为四方"
    pyramid["note"] = "备注"

    swindle["platform_name"] = "平台名称"
    swindle["platform_website"] = "平台网址"
    swindle["address_effect"] = "地址作用"
    swindle["address_type"] = "地址类型"
    swindle["sifang"] = "是否为四方"
    swindle["note"] = "备注"

    steal["note"] = "备注"

    hacker["note"] = "备注"

    square["platform_name"] = "平台名称"
    square["platform_website"] = "平台网址"
    square["address_effect"] = "地址作用"
    square["address_type"] = "地址类型"
    square["sifang"] = "是否为四方"
    square["note"] = "备注"

    mix_currency_contract["note"] = "备注"

    exception_centralization_address["note"] = "备注"


def build_header(property_name_dict, rows, max_column):
    """
    构建新表头: 第一行所有+第二行的'属性'解析
    :param property_name_dict:
    :param rows:
    :param max_column:
    :return:
    """
    header = list()
    header_index = dict()

    row_header = rows[0]
    row_1 = rows[1]

    index = 1
    # 第一行(不包含最后列'属性')
    for x in row_header:
        if x.value != "属性":
            header.append(x.value)
            header_index[x.value] = index
            index += 1

    # 第二行, 解析'属性'
    property_json = row_1[max_column - 1].value
    for x in json.loads(property_json).keys():
        header.append(property_name_dict.get(x))
        header_index[property_name_dict.get(x)] = index
        index += 1

    return header, header_index


def parse(source_file_name, to_file_name, property_name_dict):
    print("Start parse " + source_file_name)
    # source文件
    source_wb = openpyxl.load_workbook(source_file_name)
    source_ws = source_wb.active
    max_row = source_ws.max_row
    max_column = source_ws.max_column

    # to文件
    to_wb = openpyxl.Workbook()
    to_ws = to_wb["Sheet"]

    # 构建输出文件表头
    new_row_index = 1
    new_header, header_index = build_header(property_name_dict, source_ws[1:2], max_column)
    for i in range(0, len(new_header)):
        to_ws.cell(row=new_row_index, column=i + 1, value=new_header[i])
    new_row_index += 1

    for row in source_ws.rows:
        for cell in row:
            if cell.row >= 2:
                if cell.column < max_column:
                    to_ws.cell(row=new_row_index, column=cell.column, value=cell.value)
                else:
                    if cell.value is None:
                        raise Exception("value is None, row=" + cell.row + ", column=" + cell.column)
                    property_value_dict = json.loads(cell.value)
                    for k, v in property_value_dict.items():
                        to_ws.cell(row=new_row_index, column=header_index[property_name_dict[k]], value=v)
            else:
                # 必须, 否则new_row_index被多+1
                new_row_index -= 1
                break
        new_row_index += 1

    to_wb.save(to_file_name)
    to_wb.close()
    source_wb.close()
    print("Success " + to_file_name)


if __name__ == '__main__':
    init()

    now = datetime.datetime.now()
    time_export = str(now.year) + "-" + str(now.month) + "-" + str(now.day)

    start = time.time()

    # parse(source_file_name="../data/source/export_tag/暗网.xlsx",
    #       to_file_name="../data/output/export_tag/暗网_" + time_export + ".xlsx",
    #       property_name_dict=dark_net)
    #
    # parse(source_file_name="../data/source/export_tag/毒品-黑.xlsx",
    #       to_file_name="../data/output/export_tag/毒品-黑_" + time_export + ".xlsx",
    #       property_name_dict=drug)
    #
    # parse(source_file_name="../data/source/export_tag/毒品-灰.xlsx",
    #       to_file_name="../data/output/export_tag/毒品-灰_" + time_export + ".xlsx",
    #       property_name_dict=drug)
    #
    # parse(source_file_name="../data/source/export_tag/赌博-黑.xlsx",
    #       to_file_name="../data/output/export_tag/赌博-黑_" + time_export + ".xlsx",
    #       property_name_dict=gamble)
    #
    # parse(source_file_name="../data/source/export_tag/赌博-灰.xlsx",
    #       to_file_name="../data/output/export_tag/赌博-灰_" + time_export + ".xlsx",
    #       property_name_dict=gamble)
    #
    # parse(source_file_name="../data/source/export_tag/恶意地址.xlsx",
    #       to_file_name="../data/output/export_tag/恶意地址_" + time_export + ".xlsx",
    #       property_name_dict=malicious_address)

    # parse(source_file_name="../data/source/export_tag/洗钱.xlsx",
    #       to_file_name="../data/output/export_tag/洗钱_" + time_export + ".xlsx",
    #       property_name_dict=money_laundering)

    parse(source_file_name="../data/source/export_tag/钓鱼诈骗.xlsx",
          to_file_name="../data/output/export_tag/钓鱼诈骗_" + time_export + ".xlsx",
          property_name_dict=phishing_fraud)

    parse(source_file_name="../data/source/export_tag/色情.xlsx",
          to_file_name="../data/output/export_tag/色情_" + time_export + ".xlsx",
          property_name_dict=pornography)

    parse(source_file_name="../data/source/export_tag/传销-黑.xlsx",
          to_file_name="../data/output/export_tag/传销-黑_" + time_export + ".xlsx",
          property_name_dict=pyramid)

    parse(source_file_name="../data/source/export_tag/诈骗-黑.xlsx",
          to_file_name="../data/output/export_tag/诈骗-黑_" + time_export + ".xlsx",
          property_name_dict=swindle)

    end = time.time()
    print("运行时间: " + str(end - start))
